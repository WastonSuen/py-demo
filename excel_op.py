# -*- coding: utf-8 -*-

import collections
import copy
import openpyxl
import os

DEFAULT_SHEET = 'Sheet'


class WorkBookFile(object):
    def __init__(self, filepath, pr_key, read_only=False, *args, **kwargs):
        self.filepath = filepath
        self.pr_key = pr_key
        self.filedata = self._read_file(filepath, read_only)
        self.data_keys = self.get_data_keys()
        self.key_index = self.get_key_index()

        self._pre_handle_data()
        self.data = self._format_filedata_2_dict()

    def _read_file(self, filepath, read_only=False):
        filedata = openpyxl.load_workbook(filepath, read_only)
        return filedata

    @classmethod
    def load_data_from_file(cls, filepath, pr_key, read_only=False):
        u"""
        读取文件，返回类实例.
        :param filepath:
        :param pr_key: 作为唯一表示的键名称
        :param read_only:
        :return:
        """
        return cls(filepath, pr_key, read_only)

    def save(self):
        self._format_dict_2_filedata()
        self._save_to_file(self.filepath)

    def save_as(self, filepath):
        self._format_dict_2_filedata()
        self._save_to_file(filepath)

    def _save_to_file(self, filepath):
        self.filedata.save(filepath)

    @classmethod
    def new_file(cls, filepath, data_keys, *args, **kwargs):
        if os.path.exists(filepath):
            raise ValueError("File<%s> already exists" % filepath)

        pr_key = kwargs.pop('pr_key')
        wb = openpyxl.Workbook()
        wb.remove(wb[DEFAULT_SHEET])
        sheet = wb.create_sheet('Sheet1')
        sheet.append(data_keys)
        wb.save(filepath)
        return cls(filepath, pr_key)

    def get_data_keys(self):
        sheets_header = collections.OrderedDict()
        for sheet in self.filedata.worksheets:
            for row in sheet.values:
                if self.pr_key in row:
                    # 第一行为header
                    sheets_header[sheet.title] = row
                    break
            # Sheet可能为空
            else:
                sheets_header[sheet.title] = tuple()
        return sheets_header

    def get_key_index(self):
        u"""获取索引键在所有键的列表中的位置."""
        _sheet = self.filedata.worksheets[0]
        for row in _sheet.values:
            key_index = row.index(self.pr_key)
            break
        else:
            raise ValueError("Key<%s> Not Found" % self.pr_key)

        return key_index

    def _pre_handle_data(self):
        self._set_merged_cells_value()

    def _set_merged_cells_value(self):
        u"""设置合并单元格所有元素的值都为最左上角的值（这不会改变filedata的merged_cells）."""
        for sheet in self.filedata.worksheets:
            for m_cell in sheet.merged_cells.ranges:
                # 全部设置为左上角的值
                value = sheet.cell(m_cell.min_row, m_cell.min_col).value
                for row in range(m_cell.min_row, m_cell.max_row + 1):
                    for col in range(m_cell.min_col, m_cell.max_col + 1):
                        sheet.cell(row, col).value = value

    def _format_filedata_2_dict(self):

        data = collections.OrderedDict()
        for sheet in self.filedata.worksheets:
            data.setdefault(
                sheet.title,
                collections.OrderedDict()
            )

            # 键可能是合并单元格，所以值采用列表格式
            sheet_data = collections.defaultdict(list)
            for row in sheet.values:
                # 跳过第一行
                if self.pr_key in row:
                    continue

                row_key = row[self.key_index]
                # 将header与行值映射为字典
                mappings = zip(self.data_keys[sheet.title], row)
                sheet_data[row_key].append(mappings)

            data[sheet.title] = sheet_data

        return data

    def _format_dict_2_filedata(self):
        new_wb = openpyxl.Workbook()

        # 创建sheet及其header
        for sheet_header, header in self.data_keys.items():
            st = new_wb.create_sheet(sheet_header)
            st.append(header)
        
        for sheet_header, sheet_data in self.data.items():
            st = new_wb[sheet_header]
            # data是元素为字典的列表
            for data in sheet_data.values():
                for row in data:
                    # 按键的顺序写入
                    new_data_tuple = [row[key] for key in
                                      self.data_keys[sheet_header]]
                    st.append(new_data_tuple)

        if not DEFAULT_SHEET in self.data:
            new_wb.remove(new_wb[DEFAULT_SHEET])

        self.filedata = new_wb

    def _validate_append_item(self, sheet='Sheet', **new_data):
        if not sheet in self.data:
            raise ValueError("Sheet<%s> Not Found!" % sheet)

        if not self.pr_key in new_data:
            raise ValueError("Data Must Contains pr_key<%s>" % self.pr_key)

    def _format_input_data(self, sheet='Sheet', **new_data):
        new_data = copy.deepcopy(new_data)

        # 填充缺省值
        for key in self.data_keys[sheet]:
            old_value = None
            new_data.setdefault(key, old_value)

        # List, Tuple 转成字符串
        for key, value in new_data.items():
            if isinstance(value, (list, tuple)):
                value = ','.join(value)
                new_data[key] = value
        return new_data

    def _validate_update_item(self, sheet='Sheet', **new_data):
        if not sheet in self.data:
            raise ValueError("Sheet<%s> Not Found!" % sheet)

        if not self.pr_key in new_data:
            raise ValueError("Data Must Contains pr_key<%s>" % self.pr_key)

        if new_data[self.pr_key] not in self.data[sheet]:
            raise ValueError(
                "Primary<%s> Key Not Found!" % new_data[self.pr_key]
            )

    def append_single_item(self, sheet='Sheet', **new_data):
        # TODO 支持新增sheet
        # 验证值
        self._validate_append_item(sheet, **new_data)

        # 格式化值
        new_data = self._format_input_data(sheet, **new_data)

        # 增加到data dict
        primary_key = new_data[self.pr_key]
        self.data[sheet][primary_key].append(new_data)

    def update_single_item(self, sheet='Sheet', **new_data):
        u"""覆盖修改，之前的所有数据都会被覆盖."""
        self._validate_update_item(sheet, **new_data)
        new_data = self._format_input_data(sheet, **new_data)

        # 先删除该key，再添加到data dict
        primary_key = new_data[self.pr_key]
        self.data[sheet].pop(primary_key, None)
        self.data[sheet][primary_key].append(new_data)

    def get_single_item(self, index_value):
        for sheet_data in self.data.values():
            if index_value in sheet_data:
                return sheet_data[index_value]
        return None

    def delete_single_item(self, index_value):
        for sheet_data in self.data.values():
            if index_value in sheet_data:
                sheet_data.pop(index_value)
                break
        else:
            raise ValueError("Key<%s> Not Found" % index_value)


def test_for_workbook():
    if os.path.exists('testcase1.xlsx'):
        os.remove('testcase1.xlsx')

    # 创建
    data_keys = ["TestCase Class UUID", "API TestCase UUID", "API Purpose",
                 "Request URL", "Request Method", "Is Pre Step",
                 "Request Data Type", "Request Data", "Response Status",
                 "Check Point", "Active"]

    workbook = WorkBookFile.new_file(
        'testcase1.xlsx', data_keys, pr_key='API TestCase UUID'
    )
    append_data = {
        "API TestCase UUID": '066289e9-3748-4e36-bcdd-18e4db6916a7',
        "Request URL": "URL AAA"
    }
    workbook.append_single_item('Sheet1', **append_data)

    # 获取
    # workbook = WorkBookCase.load_data_from_file(
    #     filepath='testcase.xlsx', pr_key='API TestCase UUID', read_only=False
    # )
    print workbook.get_single_item('066289e9-3748-4e36-bcdd-18e4db6916a7')

    # 新增
    new_data = {
        'API TestCase UUID': 'dbe25d5f-90d9-49d2-820c-be977fab6a24',
        "Request URL": "URL BBB",
        'Request Method': 'POST',
        'Active': 'Yes'
    }
    workbook.append_single_item(sheet='Sheet1', **new_data)
    print workbook.get_single_item('dbe25d5f-90d9-49d2-820c-be977fab6a24')

    # 更新
    update_data = {
        'API TestCase UUID': 'dbe25d5f-90d9-49d2-820c-be977fab6a24',
        "Request URL": "URL CCC",
        'Request Method': 'POST',
        'Active': 'No'
    }
    workbook.append_single_item(sheet='Sheet1', **update_data)

    workbook.save_as('testcase2.xlsx')

    # 更新
    workbook.update_single_item(sheet='Sheet1', **update_data)
    print workbook.get_single_item('dbe25d5f-90d9-49d2-820c-be977fab6a24')

    workbook.save_as('testcase3.xlsx')

    # 删除
    workbook.delete_single_item("066289e9-3748-4e36-bcdd-18e4db6916a7")

    workbook.save_as('testcase4.xlsx')


if __name__ == '__main__':
    test_for_workbook()
